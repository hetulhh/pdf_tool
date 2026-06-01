"""
PDF to Excel Converter — Flask/Render edition
- API Key authentication via X-API-Key header
- Smart table filtering: ignores single-column fragment tables
- All tables merged into one sheet per page
"""

import base64
import hmac
import io
import json
import logging
import os
import re
import secrets

from flask import Flask, request
import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────
# API Key authentication
# ─────────────────────────────────────────────

_API_KEY: str = os.environ.get("PDF_API_KEY", "")


def _check_api_key() -> bool:
    if not _API_KEY:
        logger.error("PDF_API_KEY environment variable is not set. All requests blocked.")
        return False
    incoming = request.headers.get("X-API-Key", "")
    return hmac.compare_digest(
        incoming.encode("utf-8"),
        _API_KEY.encode("utf-8"),
    )


# ─────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────

@app.route("/", methods=["POST", "OPTIONS"])
def pdf_to_excel():
    """Main endpoint — returns raw .xlsx binary directly."""

    if request.method == "OPTIONS":
        return _cors_response("", 204)

    if not _check_api_key():
        return _cors_response(
            json.dumps({"error": "Unauthorized", "detail": "Invalid or missing API key"}),
            401,
            {"Content-Type": "application/json",
             "WWW-Authenticate": 'ApiKey realm="PDF Converter"'},
        )

    try:
        pdf_bytes = _extract_pdf_bytes()
    except ValueError as e:
        return _cors_response(json.dumps({"error": str(e)}), 400,
                              {"Content-Type": "application/json"})

    try:
        xlsx_bytes, summary = convert_pdf_to_excel(pdf_bytes)
    except Exception as e:
        logger.exception("Conversion failed")
        return _cors_response(json.dumps({"error": f"Conversion failed: {e}"}), 500,
                              {"Content-Type": "application/json"})

    logger.info("Conversion successful: %s", summary)

    headers = {
        "Content-Type":        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "Content-Disposition": 'attachment; filename="converted.xlsx"',
        "X-Pages":             str(summary["pages"]),
        "X-Tables-Found":      str(summary["tables_found"]),
        "X-Text-Sheets":       str(summary["text_sheets"]),
    }
    return _cors_response(xlsx_bytes, 200, headers)


# ─────────────────────────────────────────────
# PDF → Excel conversion logic
# ─────────────────────────────────────────────

# Minimum number of columns a table must have to be considered "real".
# Single-column results are usually PDF column fragments, not real tables.
MIN_TABLE_COLUMNS = 2

# If a multi-column table has ALL its data crammed into newline-separated
# cells (pdfplumber joins split columns), split them back out.
MAX_CELL_NEWLINES = 5


def convert_pdf_to_excel(pdf_bytes: bytes) -> tuple[bytes, dict]:
    summary = {"pages": 0, "tables_found": 0, "text_sheets": 0}

    wb = Workbook()
    wb.remove(wb.active)
    toc_ws = wb.create_sheet("📋 Summary", 0)

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        summary["pages"] = len(pdf.pages)
        all_table_data = []   # [(page_num, label, df)]
        all_text_data  = []   # [(page_num, text)]

        for page_num, page in enumerate(pdf.pages, start=1):

            # ── Extract tables with tighter settings ────────────────
            tables = page.extract_tables({
                "vertical_strategy":   "lines",
                "horizontal_strategy": "lines",
            })

            # Fallback: if no tables found with lines, try text strategy
            if not tables:
                tables = page.extract_tables({
                    "vertical_strategy":   "text",
                    "horizontal_strategy": "text",
                })

            real_tables = []
            for raw in tables:
                df = _clean_table(raw)
                if df is not None and len(df.columns) >= MIN_TABLE_COLUMNS:
                    df = _expand_newline_cells(df)
                    real_tables.append(df)

            # Deduplicate: remove tables that are pure subsets of a larger one
            real_tables = _deduplicate_tables(real_tables)

            for t_idx, df in enumerate(real_tables, start=1):
                label = f"P{page_num}_Table{t_idx}" if len(real_tables) > 1 else f"P{page_num}_Tables"
                all_table_data.append((page_num, label, df))

            # ── Text ────────────────────────────────────────────────
            text = (page.extract_text(x_tolerance=3, y_tolerance=3) or "").strip()
            if text:
                all_text_data.append((page_num, text))

        # ── Write sheets ────────────────────────────────────────────
        for page_num, label, df in all_table_data:
            ws = wb.create_sheet(_safe_sheet_name(label))
            _write_dataframe(ws, df, label.replace("_", " — "))
            summary["tables_found"] += 1

        for page_num, text in all_text_data:
            ws = wb.create_sheet(_safe_sheet_name(f"P{page_num}_Text"))
            _write_text(ws, text, f"Page {page_num} — Text Content")
            summary["text_sheets"] += 1

        _write_summary(toc_ws, summary, wb.sheetnames[1:])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), summary


# ─────────────────────────────────────────────
# Table helpers
# ─────────────────────────────────────────────

def _clean_table(raw_table: list) -> "pd.DataFrame | None":
    if not raw_table or len(raw_table) < 2:
        return None

    cleaned = [[str(c).strip() if c is not None else "" for c in row]
               for row in raw_table]

    header_idx = 0
    for i, row in enumerate(cleaned):
        if any(cell for cell in row):
            header_idx = i
            break

    header = cleaned[header_idx]
    rows   = cleaned[header_idx + 1:]

    seen: dict[str, int] = {}
    deduped = []
    for col in header:
        base = col if col else f"Col_{len(deduped)+1}"
        if base in seen:
            seen[base] += 1
            deduped.append(f"{base}_{seen[base]}")
        else:
            seen[base] = 0
            deduped.append(base)

    if not rows:
        return None

    df = pd.DataFrame(rows, columns=deduped)
    df.replace("", pd.NA, inplace=True)
    df.dropna(how="all", inplace=True)
    df.dropna(axis=1, how="all", inplace=True)
    df.fillna("", inplace=True)
    return df if not df.empty else None


def _expand_newline_cells(df: pd.DataFrame) -> pd.DataFrame:
    needs_expand = False
    for col in df.columns:
        newline_count = df[col].astype(str).str.contains("\n").sum()
        if newline_count > len(df) * 0.4:
            needs_expand = True
            break

    if not needs_expand:
        return df

    split_cols = {col: df[col].astype(str).str.split("\n") for col in df.columns}
    split_df = pd.DataFrame(split_cols)

    max_len = split_df.map(len).max(axis=1)
    for col in split_df.columns:
        split_df[col] = split_df.apply(
            lambda row: row[col] + [""] * (max_len[row.name] - len(row[col])),
            axis=1
        )

    rows = []
    for _, row in split_df.iterrows():
        length = max(len(v) for v in row)
        for i in range(length):
            rows.append({col: (row[col][i] if i < len(row[col]) else "") for col in df.columns})

    result = pd.DataFrame(rows, columns=df.columns)
    result.replace("", pd.NA, inplace=True)
    result.dropna(how="all", inplace=True)
    result.fillna("", inplace=True)
    return result


def _deduplicate_tables(tables: list) -> list:
    if len(tables) <= 1:
        return tables

    keep = []
    for i, t in enumerate(tables):
        t_cols = set(t.columns)
        is_subset = False
        for j, other in enumerate(tables):
            if i == j:
                continue
            other_cols = set(other.columns)
            if t_cols.issubset(other_cols) and len(t_cols) < len(other_cols):
                is_subset = True
                break
        if not is_subset:
            keep.append(t)
    return keep


# ─────────────────────────────────────────────
# Excel writing helpers
# ─────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", start_color="2F5496")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
DATA_FONT   = Font(name="Calibri", size=10)
TITLE_FONT  = Font(bold=True, name="Calibri", size=13, color="1F3864")
ALT_FILL    = PatternFill("solid", start_color="EEF2FF")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _write_dataframe(ws, df: pd.DataFrame, title: str):
    ws.append([title])
    ws["A1"].font = TITLE_FONT
    ws.append([])

    ws.append(list(df.columns))
    header_row = ws.max_row
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border    = THIN_BORDER

    for row_num, (_, row) in enumerate(df.iterrows()):
        ws.append(list(row))
        data_row = ws.max_row
        fill = ALT_FILL if row_num % 2 == 0 else None
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=data_row, column=col_idx)
            cell.font      = DATA_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(wrap_text=False)
            if fill:
                cell.fill = fill

    _autofit_columns(ws)
    ws.freeze_panes = f"A{header_row + 1}"


def _write_text(ws, text: str, title: str):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.append([])
    for line in text.split("\n"):
        ws.append([line])
        ws.cell(ws.max_row, 1).font = DATA_FONT
    ws.column_dimensions["A"].width = 120
    ws.sheet_view.showGridLines = False


def _write_summary(ws, summary: dict, sheet_names: list):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50

    ws["A1"] = "PDF → Excel Conversion Report"
    ws["A1"].font = Font(bold=True, name="Calibri", size=16, color="1F3864")
    ws.merge_cells("A1:B1")
    ws.append([])

    for label, value in [
        ("Total Pages",  summary["pages"]),
        ("Tables Found", summary["tables_found"]),
        ("Text Sheets",  summary["text_sheets"]),
        ("Total Sheets", len(sheet_names)),
    ]:
        row = ws.max_row + 1
        ws.cell(row, 1, label).font = Font(bold=True, name="Calibri", size=11)
        ws.cell(row, 2, value).font = Font(name="Calibri", size=11)

    ws.append([])
    ws.append(["Sheet", "Description"])
    hrow = ws.max_row
    for col in (1, 2):
        c = ws.cell(hrow, col)
        c.fill      = HEADER_FILL
        c.font      = HEADER_FONT
        c.alignment = Alignment(horizontal="center")

    for name in sheet_names:
        ws.append([name, _sheet_description(name)])
        for col in (1, 2):
            ws.cell(ws.max_row, col).font = DATA_FONT


def _autofit_columns(ws, max_width=40):
    for col in ws.columns:
        max_len    = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val.split("\n")[0]))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, max_width)


def _safe_sheet_name(name: str) -> str:
    return re.sub(r"[\\/*?:\[\]]", "_", name)[:31]


def _sheet_description(name: str) -> str:
    if "Table" in name:
        return "Extracted table data"
    if "Text" in name:
        return "Extracted text content"
    return "Data sheet"


# ─────────────────────────────────────────────
# Request parsing
# ─────────────────────────────────────────────

def _extract_pdf_bytes() -> bytes:
    content_type = request.content_type or ""

    if "multipart/form-data" in content_type:
        file = request.files.get("file")
        if not file:
            raise ValueError("No file field found in multipart form data.")
        data = file.read()
        if not data:
            raise ValueError("Uploaded file is empty.")
        return data

    if "application/json" in content_type:
        body = request.get_json(silent=True) or {}
        b64  = body.get("pdf_base64") or body.get("file")
        if not b64:
            raise ValueError("JSON body must include 'pdf_base64' key.")
        try:
            return base64.b64decode(b64)
        except Exception:
            raise ValueError("'pdf_base64' is not valid base64.")

    data = request.get_data()
    if data:
        return data

    raise ValueError(
        "Send PDF as: multipart/form-data (field='file'), "
        "JSON {'pdf_base64':'<b64>'}, or raw binary body."
    )


def _cors_response(body, status=200, extra_headers=None):
    headers = {
        "Access-Control-Allow-Origin":  "*",
        "Access-Control-Allow-Methods": "POST, OPTIONS",
        "Access-Control-Allow-Headers": "Content-Type, X-API-Key",
    }
    if extra_headers:
        headers.update(extra_headers)
    return body, status, headers


# ─────────────────────────────────────────────
# Key generation utility
# ─────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == "genkey":
        key = secrets.token_urlsafe(32)
        print(f"\nGenerated API key:\n\n  {key}\n")
        print("Set on Render dashboard as:")
        print(f"  PDF_API_KEY = {key}\n")
    else:
        app.run(host="0.0.0.0", port=10000, debug=False)
